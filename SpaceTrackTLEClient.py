import configparser
import requests

from pathlib import Path
from datetime import datetime
from urllib.parse import quote
from TLEParser import TLE

from openpyxl import Workbook, load_workbook

class SpaceTrackTLEClient:
    URI_BASE = "https://www.space-track.org"
    LOGIN_PATH = "/ajaxauth/login"
    QUERY_PATH = "/basicspacedata/query"

    def __init__(
        self,
        ini_path: str = "./SLTrack.ini",
        timeout_s: int = 30,
        user_agent: str = "SpaceTrackTLEClient/1.0",
    ):
        config = configparser.ConfigParser()
        if not config.read(ini_path):
            raise FileNotFoundError(f"INI not found or unreadable: {ini_path}")

        self.username = config.get("configuration", "username")
        self.password = config.get("configuration", "password")
        self.excel_path = Path(config.get("configuration", "output")).expanduser()

        self.timeout_s = timeout_s
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": user_agent})
        self.logged_in = False

    # -----------------------------
    # Auth
    # -----------------------------
    def login(self) -> None:
        url = self.URI_BASE + self.LOGIN_PATH
        response = self.session.post(
            url,
            data={"identity": self.username, "password": self.password},
            timeout=self.timeout_s,
        )
        if response.status_code != 200:
            raise RuntimeError(f"Space-Track login failed HTTP {response.status_code}: {response.text[:300]}")
        self.logged_in = True

        # Check that session cookie exists
        if "chocolatechip" not in self.session.cookies.get_dict():
            raise RuntimeError("Login failed: session cookie not received.")

    def ensure_login(self) -> None:
        if not self.logged_in:
            self.login()

    # -----------------------------
    # Query helpers
    # -----------------------------
    @staticmethod
    def _url_escape(value) -> str:
        return quote(str(value), safe="")

    def build_query_url(self, class_name: str, **segments) -> str:
        """
        Build a Space-Track REST-style query URL.
        Example:
        /basicspacedata/query/class/gp/NORAD_CAT_ID/25544/format/tle
        """
        url_parts = [self.URI_BASE + self.QUERY_PATH, "class", class_name]
        for key, value in segments.items():
            url_parts.extend([key, self._url_escape(value)])

        return "/".join(url_parts)

    def fetch_text(self, url: str) -> str:
        """
        Fetch the response body as text.
        """
        self.ensure_login()
        response = self.session.get(url, timeout=self.timeout_s)
        if response.status_code != 200:
            raise RuntimeError(f"Space-Track query failed HTTP {response.status_code}: {response.text[:300]}")
        
        return response.text

    # -----------------------------
    # Excel helpers
    # -----------------------------
    def _open_or_create_workbook(self):
        """
        Open existing workbook or create a new one.
        Ensures parent directory exists.
        """
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        if self.excel_path.exists():
            return load_workbook(self.excel_path)
        
        return Workbook()

    @staticmethod
    def _ensure_sheet(workbook, name: str, headers: list[str]):
        """
        Return worksheet. If it does not exist, create it and write headers.
        """
        if name in workbook.sheetnames:
            worksheet = workbook[name]
            # Only write headers if sheet is empty
            if worksheet["A1"].value is None:
                worksheet.append(headers)

            return worksheet
        
        worksheet = workbook.create_sheet(title=name)
        worksheet.append(headers)

        return worksheet

    @staticmethod
    def _remove_default_sheet_if_empty(workbook):
        """
        Remove the default 'Sheet' if it is unused.
        """
        default_name = "Sheet"

        if default_name in workbook.sheetnames:
            default_sheet = workbook[default_name]

            is_empty = (
                default_sheet.max_row == 1
                and default_sheet.max_column == 1
                and default_sheet["A1"].value is None
            )

            if is_empty:
                workbook.remove(default_sheet)

    # -----------------------------
    # Main method: fetch 1 NORAD and save to Excel
    # -----------------------------
    def fetch_latest_by_norad_and_save(
        self,
        norad_id: int,
        sheet_name: str = "tle_by_norad",
        save_lines: bool = True,
    ) -> tuple[TLE, Path]:
        """
        Fetch latest TLE for a single NORAD ID, append one row to Excel,
        and return (tle_obj, excel_path).
        """
        query_url = self.build_query_url(
            "gp",
            NORAD_CAT_ID=norad_id,
            orderby="EPOCH desc",
            limit=1,
            format="tle",
        )

        tle_text = self.fetch_text(query_url)
        tle_obj = TLE.from_text(tle_text)

        workbook = self._open_or_create_workbook()

        headers = [
            "timestamp",
            "name", "sat_num", "elset_class", "int_designator",
            "epoch_utc", "epoch_year", "epoch_day",
            "n_dot", "n_ddot", "bstar",
            "elem_set_type", "elem_number",
            "inc", "raan", "ecc", "argp", "M", "n", "rev_num",
        ]
        if save_lines:
            headers += ["line1", "line2", "query_url"]

        worksheet = self._ensure_sheet(workbook, sheet_name, headers=headers)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        row = [
            timestamp,
            tle_obj.name, tle_obj.sat_num, tle_obj.elset_class, tle_obj.int_designator,
            tle_obj.epoch_utc, tle_obj.epoch_year, tle_obj.epoch_day,
            tle_obj.n_dot, tle_obj.n_ddot, tle_obj.bstar,
            tle_obj.elem_set_type, tle_obj.elem_number,
            tle_obj.inc, tle_obj.raan, tle_obj.ecc, tle_obj.argp, tle_obj.M, tle_obj.n, tle_obj.rev_num,
        ]
        if save_lines:
            row += [tle_obj.line1, tle_obj.line2, query_url]

        worksheet.append(row)

        self._remove_default_sheet_if_empty(workbook)
        workbook.save(self.excel_path)

        return tle_obj, self.excel_path